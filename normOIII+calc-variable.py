from __future__ import annotations

import warnings
from pathlib import Path
from typing import Optional

import matplotlib
matplotlib.use("Agg")

import matplotlib.dates as mdates
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from astropy.io import fits
from astropy.time import Time
from scipy.ndimage import gaussian_filter1d
from scipy.optimize import curve_fit
from scipy.stats import chi2 as chi2_dist

try:
    from openpyxl.utils import get_column_letter
except Exception:
    get_column_letter = None

warnings.filterwarnings("ignore", category=RuntimeWarning)


# ============================================================
# НАСТРОЙКИ
# ============================================================

ROOT_DIR = Path(r"C:\Users\IvanK\Astro\Course5\downloaded_spectra")

INPUT_MERGED_TABLE = "line_and_continuum_vs_time.tsv"
INPUT_LINE_TABLE = "line_fluxes_vs_time.tsv"
INPUT_CONT_TABLE = "continuum_5100_vs_time.tsv"

OIII_NORM_TABLE = "oiii5007_normalization.tsv"
OUTPUT_NORM_TABLE = "line_and_continuum_vs_time_normalized.tsv"
OUTPUT_VAR_TABLE = "variability_stats.tsv"
OUTPUT_PLOT = "variability_diagnostic.png"

MASTER_VAR_TABLE_TSV = "_variability_stats_all_objects.tsv"
MASTER_VAR_TABLE_XLSX = "_variability_stats_all_objects.xlsx"

MASTER_OBJ_TABLE_TSV = "_rm_candidate_summary.tsv"
MASTER_OBJ_TABLE_XLSX = "_rm_candidate_summary.xlsx"

REDSHIFT_CACHE_NAME = "_redshift_cache.tsv"

# [O III] 5007
OIII_5007 = 5006.843
OIII_4959 = 4958.911
OIII_RATIO_5007_TO_4959 = 2.98

# окно фита [O III]
OIII_FIT_MIN = 4935.0
OIII_FIT_MAX = 5038.0

# окно прямого интегрирования (fallback)
OIII_INT_MIN = 4998.0
OIII_INT_MAX = 5016.0
OIII_CONT1_MIN = 4972.0
OIII_CONT1_MAX = 4988.0
OIII_CONT2_MIN = 5020.0
OIII_CONT2_MAX = 5036.0

# сглаживание перед измерением [O III]
USE_OIII_SMOOTHING = True
OIII_SMOOTH_SIGMA_PIX = 1.5

# минимальный S/N [O III], чтобы использовать спектр в референсе
MIN_OIII_SNR_FOR_REFERENCE = 3.0

# критерии переменности
P_CONST_THRESHOLD = 0.01
PAIRWISE_SIG_THRESHOLD = 3.0
PAIRWISE_SIG_STRONG = 5.0
FVAR_SIGMA_THRESHOLD = 3.0

# стили графиков
SURVEY_STYLES = {
    "SDSS": {"color": "royalblue", "marker": "o"},
    "DESI": {"color": "darkorange", "marker": "s"},
    "OTHER": {"color": "gray", "marker": "^"},
}

# Предпочтительный порядок столбцов
PREFERRED_VAR_COLS = [
    "object_name",
    "series",
    "n_epochs",
    "mean_flux",
    "weighted_mean_flux",
    "chi2_const",
    "chi2_red",
    "p_const",
    "sigma_xs2",
    "sigma_nxs2",
    "fvar",
    "fvar_err",
    "rmax",
    "max_pair_sig",
    "variable",
    "variable_loose",
    "classification",
    "criterion",
]

PREFERRED_SUMMARY_COLS = [
    "object_name",
    "rm_candidate",
    "cont_variable",
    "ha_variable",
    "hb_variable",
    "n_spectra_total",
    "n_oiii_measured",
    "oiii_ref_flux",
    "cont_n",
    "cont_p_const",
    "cont_fvar",
    "cont_fvar_err",
    "ha_n",
    "ha_p_const",
    "ha_fvar",
    "ha_fvar_err",
    "hb_n",
    "hb_p_const",
    "hb_fvar",
    "hb_fvar_err",
    "oiii_raw_p_const",
    "oiii_raw_fvar",
]

PREFERRED_OIII_COLS = [
    "object_name",
    "file_name",
    "survey",
    "date_obs",
    "date_mjd",
    "redshift",
    "redshift_source",
    "specid",
    "targetid",
    "oiii5007_flux",
    "oiii5007_err",
    "oiii5007_snr",
    "oiii_method",
    "status",
    "notes",
    "spectrum_file",
    "oiii_plot_file",
]

PREFERRED_NORM_COLS = [
    "object_name",
    "file_name",
    "survey",
    "date_obs",
    "date_mjd",
    "redshift",
    "specid",
    "targetid",
    "oiii5007_flux",
    "oiii5007_err",
    "oiii5007_snr",
    "oiii_ref_flux",
    "oiii_scale_factor",
    "ha_broad_flux",
    "ha_broad_err",
    "ha_broad_flux_norm",
    "ha_broad_err_norm",
    "hb_broad_flux",
    "hb_broad_err",
    "hb_broad_flux_norm",
    "hb_broad_err_norm",
    "cont5100_int_flux",
    "cont5100_err",
    "cont5100_int_flux_norm",
    "cont5100_err_norm",
    "spectrum_file",
]


# ============================================================
# СОХРАНЕНИЕ ТАБЛИЦ В ЧИТАЕМОМ ВИДЕ
# ============================================================

def reorder_columns(df: pd.DataFrame, preferred_cols: list[str]) -> pd.DataFrame:
    if df.empty:
        return df.copy()

    cols_main = [c for c in preferred_cols if c in df.columns]
    cols_rest = [c for c in df.columns if c not in cols_main]
    return df[cols_main + cols_rest].copy()


def save_tsv(
    df: pd.DataFrame,
    path: Path,
    preferred_cols: Optional[list[str]] = None,
):
    out = df.copy()
    if preferred_cols:
        out = reorder_columns(out, preferred_cols)

    out.to_csv(
        path,
        sep="\t",
        index=False,
        encoding="utf-8-sig",
        na_rep="NA",
        float_format="%.6e",
    )


def save_table_readable(
    df: pd.DataFrame,
    tsv_path: Path,
    xlsx_path: Optional[Path] = None,
    preferred_cols: Optional[list[str]] = None,
):
    out = df.copy()
    if preferred_cols:
        out = reorder_columns(out, preferred_cols)

    save_tsv(out, tsv_path)

    if xlsx_path is not None:
        try:
            with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
                out.to_excel(writer, index=False, sheet_name="data")

                ws = writer.sheets["data"]
                ws.freeze_panes = "A2"

                if get_column_letter is not None:
                    for i, col in enumerate(out.columns, start=1):
                        values = [str(col)]
                        values.extend([str(v) if pd.notna(v) else "NA" for v in out[col].values])
                        max_len = min(max(len(v) for v in values) + 2, 40)
                        ws.column_dimensions[get_column_letter(i)].width = max_len
        except Exception as e:
            print(f"[WARN] Не удалось сохранить Excel-файл {xlsx_path.name}: {e}")


# ============================================================
# ОБЩИЕ ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ
# ============================================================

def safe_float(value) -> float:
    try:
        if value is None or pd.isna(value):
            return np.nan
        return float(value)
    except Exception:
        return np.nan


def safe_int(value) -> Optional[int]:
    try:
        if value is None or pd.isna(value):
            return None
        return int(value)
    except Exception:
        try:
            return int(float(value))
        except Exception:
            return None


def parse_date_to_mjd(date_str: str) -> float:
    if not date_str:
        return np.nan

    s = str(date_str).strip()
    if not s or s.lower() == "nan":
        return np.nan

    candidates = [
        s,
        s.replace(" ", "T"),
        s.replace("+00", ""),
        s.replace(" ", "T").replace("+00", ""),
    ]

    for c in candidates:
        try:
            return float(Time(c).mjd)
        except Exception:
            pass

    return np.nan


def mjd_to_datetime(mjd: float):
    return Time(float(mjd), format="mjd").to_datetime()


def strip_fits_suffix(path: Path) -> str:
    name = path.name
    low = name.lower()
    for suf in [".fits.gz", ".fit.gz", ".fits", ".fit", ".fz"]:
        if low.endswith(suf):
            return name[:-len(suf)]
    return path.stem


def list_object_dirs(root: Path) -> list[Path]:
    return sorted([p for p in root.iterdir() if p.is_dir() and p.name.startswith("J")])


def list_spectrum_files(object_dir: Path) -> list[Path]:
    files = []
    for sub in ["sdss", "desi"]:
        subdir = object_dir / sub
        if not subdir.exists():
            continue
        for pattern in ["*.fits", "*.fit", "*.fits.gz", "*.fit.gz", "*.fz"]:
            files.extend(sorted(subdir.glob(pattern)))

    uniq = []
    seen = set()
    for p in files:
        key = str(p.resolve()) if p.exists() else str(p)
        if key not in seen:
            uniq.append(p)
            seen.add(key)
    return uniq


def guess_survey_from_path(path: Path) -> str:
    parent = path.parent.name.lower()
    if parent == "sdss":
        return "SDSS"
    if parent == "desi":
        return "DESI"
    return "OTHER"


def ensure_file_name_column(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    if "file_name" not in out.columns:
        if "spectrum_file" in out.columns:
            out["file_name"] = out["spectrum_file"].apply(lambda x: Path(str(x)).name if pd.notna(x) else "")
    return out


def load_table(path: Path) -> pd.DataFrame:
    if not path.exists():
        return pd.DataFrame()
    try:
        return pd.read_csv(path, sep="\t")
    except Exception:
        return pd.DataFrame()


def first_existing_value(row: dict, keys: list[str], as_float: bool = False):
    for k in keys:
        if k in row and row[k] is not None and not pd.isna(row[k]):
            return safe_float(row[k]) if as_float else row[k]
    return np.nan if as_float else None


def load_redshift_cache(root_dir: Path) -> dict[int, dict]:
    path = root_dir / REDSHIFT_CACHE_NAME
    df = load_table(path)
    if df.empty or "specid" not in df.columns:
        return {}

    out = {}
    for _, r in df.iterrows():
        specid = safe_int(r.get("specid"))
        if specid is None:
            continue
        out[specid] = {
            "redshift": safe_float(r.get("redshift")),
            "release": str(r.get("release", "")) if r.get("release") is not None else "",
            "targetid": safe_int(r.get("targetid")),
        }
    return out


# ============================================================
# ПОДГОТОВКА ВХОДНОЙ ТАБЛИЦЫ ИЗ ВАШИХ ПРЕДЫДУЩИХ ФАЙЛОВ
# ============================================================

def combine_prefer_left(a: pd.Series, b: pd.Series) -> pd.Series:
    return a.combine_first(b)


def build_measurement_table(object_dir: Path) -> pd.DataFrame:
    merged_path = object_dir / INPUT_MERGED_TABLE
    merged_df = load_table(merged_path)
    merged_df = ensure_file_name_column(merged_df)
    if not merged_df.empty:
        return merged_df

    line_df = ensure_file_name_column(load_table(object_dir / INPUT_LINE_TABLE))
    cont_df = ensure_file_name_column(load_table(object_dir / INPUT_CONT_TABLE))

    if line_df.empty and cont_df.empty:
        return pd.DataFrame()

    if line_df.empty:
        return cont_df

    if cont_df.empty:
        return line_df

    df = pd.merge(line_df, cont_df, on="file_name", how="outer", suffixes=("_line", "_cont"))

    out = pd.DataFrame()
    out["file_name"] = df["file_name"]

    meta_cols = ["object_name", "survey", "spectrum_file", "specid", "targetid", "date_obs", "date_mjd", "redshift"]

    for col in meta_cols:
        left = f"{col}_line"
        right = f"{col}_cont"
        if left in df.columns and right in df.columns:
            out[col] = combine_prefer_left(df[left], df[right])
        elif left in df.columns:
            out[col] = df[left]
        elif right in df.columns:
            out[col] = df[right]

    for col in line_df.columns:
        if col not in out.columns and col != "file_name":
            src = f"{col}_line" if f"{col}_line" in df.columns else col
            if src in df.columns:
                out[col] = df[src]

    for col in cont_df.columns:
        if col not in out.columns and col != "file_name":
            src = f"{col}_cont" if f"{col}_cont" in df.columns else col
            if src in df.columns:
                out[col] = df[src]

    return out


# ============================================================
# ЧТЕНИЕ FITS: SDSS / DESI
# ============================================================

def preprocess_spectrum(wave: np.ndarray, flux: np.ndarray, ivar: Optional[np.ndarray] = None):
    wave = np.asarray(wave, dtype=float).copy()
    flux = np.asarray(flux, dtype=float).copy()

    mask = np.isfinite(wave) & np.isfinite(flux) & (wave > 0)

    if ivar is not None:
        ivar = np.asarray(ivar, dtype=float).copy()
        if ivar.shape == flux.shape:
            mask &= np.isfinite(ivar)
        else:
            ivar = None

    wave = wave[mask]
    flux = flux[mask]
    if ivar is not None:
        ivar = ivar[mask]

    if len(wave) < 20:
        raise ValueError("слишком мало валидных точек в спектре")

    order = np.argsort(wave)
    wave = wave[order]
    flux = flux[order]
    if ivar is not None:
        ivar = ivar[order]

    uniq = np.concatenate([[True], np.diff(wave) > 0])
    wave = wave[uniq]
    flux = flux[uniq]
    if ivar is not None:
        ivar = ivar[uniq]

    return wave, flux, ivar


def extract_date_from_hdul(hdul: fits.HDUList) -> tuple[str, float]:
    for h in hdul:
        hdr = h.header
        for key in ("DATE-OBS", "DATEOBS"):
            if hdr.get(key):
                s = str(hdr.get(key)).strip()
                return s, parse_date_to_mjd(s)

    for h in hdul:
        hdr = h.header
        for key in ("MJD", "MJD-OBS", "MJDOBS", "MEANMJD", "MINMJD", "MAXMJD"):
            val = safe_float(hdr.get(key))
            if np.isfinite(val):
                try:
                    return Time(val, format="mjd").isot, val
                except Exception:
                    return "", val

    return "", np.nan


def extract_redshift_from_hdul(hdul: fits.HDUList) -> tuple[float, str]:
    keys = ["ZUSED", "REDSHIFT", "Z", "Z_OBJ", "ZOBJ", "SPEC_Z"]
    for h in hdul:
        hdr = h.header
        for key in keys:
            val = safe_float(hdr.get(key))
            if np.isfinite(val):
                return val, f"header:{key}"
    return np.nan, ""


def read_sdss_spectrum(hdul: fits.HDUList):
    for h in hdul:
        if not isinstance(h, (fits.BinTableHDU, fits.TableHDU)) or h.data is None:
            continue
        cols = {str(c).lower(): c for c in h.columns.names}
        if "flux" in cols and "loglam" in cols:
            wave = 10 ** np.asarray(h.data[cols["loglam"]], dtype=float)
            flux = np.asarray(h.data[cols["flux"]], dtype=float)
            ivar = np.asarray(h.data[cols["ivar"]], dtype=float) if "ivar" in cols else None
            return wave, flux, ivar
    raise ValueError("не распознан SDSS-спектр")


def read_desi_spectrum(hdul: fits.HDUList):
    names = {h.name.upper(): i for i, h in enumerate(hdul)}

    if "WAVELENGTH" in names and "FLUX" in names:
        wave = np.asarray(hdul[names["WAVELENGTH"]].data, dtype=float).ravel()
        flux = np.asarray(hdul[names["FLUX"]].data, dtype=float).ravel()
        ivar = None
        if "IVAR" in names:
            ivar = np.asarray(hdul[names["IVAR"]].data, dtype=float).ravel()
        return wave, flux, ivar

    if hdul[0].data is not None:
        arr = np.asarray(hdul[0].data, dtype=float)
        if arr.ndim == 2 and arr.shape[0] >= 2:
            wave = arr[0, :]
            flux = arr[1, :]
            return wave.astype(float), flux.astype(float), None

    raise ValueError("не распознан DESI-спектр")


def load_raw_spectrum(path: Path) -> dict:
    survey_guess = guess_survey_from_path(path)

    with fits.open(path, memmap=False) as hdul:
        hdr0 = hdul[0].header

        survey = str(hdr0.get("SURVEY", survey_guess)).strip().upper() if hdr0.get("SURVEY") else survey_guess
        specid = safe_int(hdr0.get("SPECID"))
        targetid = safe_int(hdr0.get("TARGETID"))

        date_obs, date_mjd = extract_date_from_hdul(hdul)
        redshift, redshift_source = extract_redshift_from_hdul(hdul)

        if survey == "SDSS":
            wave, flux, ivar = read_sdss_spectrum(hdul)
        elif survey == "DESI":
            wave, flux, ivar = read_desi_spectrum(hdul)
        else:
            try:
                wave, flux, ivar = read_sdss_spectrum(hdul)
                survey = "SDSS"
            except Exception:
                wave, flux, ivar = read_desi_spectrum(hdul)
                survey = "DESI"

    return {
        "survey": survey,
        "specid": specid,
        "targetid": targetid,
        "date_obs": date_obs,
        "date_mjd": date_mjd,
        "redshift": redshift,
        "redshift_source": redshift_source,
        "wave_obs": wave,
        "flux": flux,
        "ivar": ivar,
    }


# ============================================================
# ИЗМЕРЕНИЕ [O III] λ5007
# ============================================================

def oiii_model_tied(x, c0, c1, amp5007, dcenter, sigma):
    cont = c0 + c1 * (x - 5000.0)
    mu5007 = OIII_5007 + dcenter
    mu4959 = OIII_4959 + dcenter

    g5007 = amp5007 * np.exp(-0.5 * ((x - mu5007) / sigma) ** 2)
    g4959 = (amp5007 / OIII_RATIO_5007_TO_4959) * np.exp(-0.5 * ((x - mu4959) / sigma) ** 2)

    return cont + g4959 + g5007


def fit_linear_continuum(x: np.ndarray, y: np.ndarray):
    if len(x) < 2:
        return np.nanmedian(y), 0.0
    p = np.polyfit(x - 5000.0, y, 1)
    return p[1], p[0]


def measure_oiii_direct(wrest: np.ndarray, flux_use: np.ndarray, ivar: Optional[np.ndarray]):
    cont_mask = ((wrest >= OIII_CONT1_MIN) & (wrest <= OIII_CONT1_MAX)) | ((wrest >= OIII_CONT2_MIN) & (wrest <= OIII_CONT2_MAX))
    line_mask = (wrest >= OIII_INT_MIN) & (wrest <= OIII_INT_MAX)

    if np.sum(cont_mask) < 4 or np.sum(line_mask) < 3:
        raise ValueError("слишком мало точек для direct OIII integration")

    c0, c1 = fit_linear_continuum(wrest[cont_mask], flux_use[cont_mask])
    cont = c0 + c1 * (wrest[line_mask] - 5000.0)
    line_flux = flux_use[line_mask] - cont

    flux_oiii = float(np.trapz(line_flux, wrest[line_mask]))

    if ivar is not None:
        iv = ivar[line_mask]
        good = np.isfinite(iv) & (iv > 0)
        if np.sum(good) >= 2:
            sigma = np.full(np.sum(line_mask), np.nan)
            sigma[good] = 1.0 / np.sqrt(iv[good])
            med = np.nanmedian(sigma[good])
            sigma = np.where(np.isfinite(sigma), sigma, med)
            dl = np.gradient(wrest[line_mask])
            err = float(np.sqrt(np.nansum((sigma * dl) ** 2)))
        else:
            resid = flux_use[cont_mask] - (c0 + c1 * (wrest[cont_mask] - 5000.0))
            rms = np.nanstd(resid)
            dl = np.gradient(wrest[line_mask])
            err = float(rms * np.sqrt(np.nansum(dl ** 2)))
    else:
        resid = flux_use[cont_mask] - (c0 + c1 * (wrest[cont_mask] - 5000.0))
        rms = np.nanstd(resid)
        dl = np.gradient(wrest[line_mask])
        err = float(rms * np.sqrt(np.nansum(dl ** 2)))

    return {
        "method": "direct",
        "flux_oiii": flux_oiii,
        "err_oiii": err,
        "c0": c0,
        "c1": c1,
        "params": None,
        "model_flux": None,
        "fit_wave": wrest[(wrest >= OIII_FIT_MIN) & (wrest <= OIII_FIT_MAX)],
    }


def measure_oiii_5007(wave_obs: np.ndarray, flux: np.ndarray, ivar: Optional[np.ndarray], redshift: float):
    wave_obs, flux, ivar = preprocess_spectrum(wave_obs, flux, ivar)

    if not np.isfinite(redshift):
        raise ValueError("не найден redshift для измерения [O III]")

    wrest = wave_obs / (1.0 + redshift)

    if USE_OIII_SMOOTHING and OIII_SMOOTH_SIGMA_PIX > 0:
        flux_use = gaussian_filter1d(flux, sigma=OIII_SMOOTH_SIGMA_PIX, mode="nearest")
    else:
        flux_use = flux.copy()

    fit_mask = (wrest >= OIII_FIT_MIN) & (wrest <= OIII_FIT_MAX)
    if np.sum(fit_mask) < 12:
        raise ValueError("слишком мало точек в окне [O III]")

    x = wrest[fit_mask]
    y = flux_use[fit_mask]

    iv_fit = None
    if ivar is not None:
        iv_fit = ivar[fit_mask]

    outer_mask = (x < 4950.0) | (x > 5018.0)
    if np.sum(outer_mask) >= 4:
        c0, c1 = fit_linear_continuum(x[outer_mask], y[outer_mask])
    else:
        c0, c1 = np.nanmedian(y), 0.0

    cont_guess = c0 + c1 * (x - 5000.0)
    amp0 = np.nanmax(y[(x > 5003) & (x < 5010)]) - np.nanmedian(cont_guess)
    if not np.isfinite(amp0) or amp0 <= 0:
        amp0 = max(np.nanpercentile(y, 97) - np.nanpercentile(y, 50), 1e-6)

    p0 = [c0, c1, amp0, 0.0, 2.0]
    bounds = (
        [-np.inf, -np.inf, 0.0, -3.0, 0.5],
        [np.inf, np.inf, np.inf, 3.0, 8.0],
    )

    try:
        if iv_fit is not None:
            good = np.isfinite(iv_fit) & (iv_fit > 0)
            if np.sum(good) >= max(5, len(iv_fit) // 2):
                sigma_fit = np.full_like(y, np.nan, dtype=float)
                sigma_fit[good] = 1.0 / np.sqrt(iv_fit[good])
                med = np.nanmedian(sigma_fit[good])
                sigma_fit = np.where(np.isfinite(sigma_fit), sigma_fit, med)

                popt, pcov = curve_fit(
                    oiii_model_tied,
                    x,
                    y,
                    p0=p0,
                    bounds=bounds,
                    sigma=sigma_fit,
                    absolute_sigma=True,
                    maxfev=100000,
                )
            else:
                popt, pcov = curve_fit(
                    oiii_model_tied,
                    x,
                    y,
                    p0=p0,
                    bounds=bounds,
                    maxfev=100000,
                )
        else:
            popt, pcov = curve_fit(
                oiii_model_tied,
                x,
                y,
                p0=p0,
                bounds=bounds,
                maxfev=100000,
            )

        c0_fit, c1_fit, amp5007, dcenter, sigma = popt
        model = oiii_model_tied(x, *popt)
        continuum = c0_fit + c1_fit * (x - 5000.0)

        flux_oiii = float(np.sqrt(2.0 * np.pi) * amp5007 * sigma)

        err_oiii = np.nan
        try:
            if pcov is not None and np.all(np.isfinite(pcov)):
                var_a = pcov[2, 2]
                var_s = pcov[4, 4]
                cov_as = pcov[2, 4]
                k = np.sqrt(2.0 * np.pi)
                var_flux = (k * sigma) ** 2 * var_a + (k * amp5007) ** 2 * var_s + 2.0 * (k ** 2) * amp5007 * sigma * cov_as
                if np.isfinite(var_flux) and var_flux >= 0:
                    err_oiii = float(np.sqrt(var_flux))
        except Exception:
            err_oiii = np.nan

        if not np.isfinite(err_oiii) or err_oiii <= 0:
            resid = y - model
            rms = np.nanstd(resid)
            err_oiii = float(np.sqrt(2.0 * np.pi) * max(sigma, 1.0) * rms)

        return {
            "method": "gaussian_tied",
            "wave_rest": wrest,
            "flux_raw": flux,
            "flux_use": flux_use,
            "fit_wave": x,
            "fit_flux": y,
            "fit_model": model,
            "fit_cont": continuum,
            "params": popt,
            "pcov": pcov,
            "flux_oiii": flux_oiii,
            "err_oiii": err_oiii,
        }

    except Exception:
        direct = measure_oiii_direct(wrest, flux_use, ivar)
        return {
            "method": direct["method"],
            "wave_rest": wrest,
            "flux_raw": flux,
            "flux_use": flux_use,
            "fit_wave": direct["fit_wave"],
            "fit_flux": flux_use[(wrest >= OIII_FIT_MIN) & (wrest <= OIII_FIT_MAX)],
            "fit_model": None,
            "fit_cont": None,
            "params": None,
            "pcov": None,
            "flux_oiii": direct["flux_oiii"],
            "err_oiii": direct["err_oiii"],
        }


def save_oiii_plot(out_path: Path, object_name: str, file_name: str, survey: str, date_obs: str, redshift: float, meas: dict):
    wave_rest = meas["wave_rest"]
    flux_raw = meas["flux_raw"]
    flux_use = meas["flux_use"]

    plot_mask = (wave_rest >= 4920) & (wave_rest <= 5050)
    if np.sum(plot_mask) < 10:
        plot_mask = np.ones_like(wave_rest, dtype=bool)

    fig, ax = plt.subplots(figsize=(10, 5.5))
    ax.plot(wave_rest[plot_mask], flux_raw[plot_mask], color="0.75", lw=1.0, label="Raw")
    ax.plot(wave_rest[plot_mask], flux_use[plot_mask], color="black", lw=1.2, label="Used for fit")

    if meas["fit_model"] is not None:
        ax.plot(meas["fit_wave"], meas["fit_model"], color="crimson", lw=2.0, label=f"Model ({meas['method']})")
    if meas["fit_cont"] is not None:
        ax.plot(meas["fit_wave"], meas["fit_cont"], color="cyan", lw=1.2, ls="--", label="Continuum")

    ax.axvline(OIII_4959, color="gray", ls="--", alpha=0.6)
    ax.axvline(OIII_5007, color="gray", ls="--", alpha=0.6)
    ax.axvspan(OIII_INT_MIN, OIII_INT_MAX, color="gold", alpha=0.15, label="Direct integration window")

    flux_val = meas["flux_oiii"]
    err_val = meas["err_oiii"]
    snr = flux_val / err_val if np.isfinite(err_val) and err_val > 0 else np.nan

    ax.set_title(
        f"{object_name} | {survey} | {file_name}\n"
        f"date={date_obs or 'NA'} | z={redshift:.6f} | "
        f"F([O III]5007)={flux_val:.3e} ± {err_val:.3e} | S/N={snr:.2f}"
    )
    ax.set_xlabel("Rest wavelength [Å]")
    ax.set_ylabel("Flux")
    ax.grid(True, alpha=0.3)
    ax.legend()
    plt.tight_layout()
    fig.savefig(out_path, dpi=150, bbox_inches="tight")
    plt.close(fig)


# ============================================================
# НОРМИРОВКА ПО [O III]
# ============================================================

def normalize_one_value(value, err_value, oiii_flux, oiii_err, ref_flux):
    value = safe_float(value)
    err_value = safe_float(err_value)
    oiii_flux = safe_float(oiii_flux)
    oiii_err = safe_float(oiii_err)
    ref_flux = safe_float(ref_flux)

    if not np.isfinite(value) or not np.isfinite(oiii_flux) or not np.isfinite(ref_flux):
        return np.nan, np.nan

    if oiii_flux <= 0 or ref_flux <= 0:
        return np.nan, np.nan

    scale = ref_flux / oiii_flux
    value_norm = value * scale

    variance = 0.0
    if np.isfinite(err_value) and err_value >= 0:
        variance += (scale * err_value) ** 2

    if np.isfinite(oiii_err) and oiii_err >= 0:
        variance += (value * ref_flux * oiii_err / (oiii_flux ** 2)) ** 2

    err_norm = np.sqrt(variance) if variance > 0 else np.nan
    return float(value_norm), float(err_norm)


def apply_oiii_normalization(df: pd.DataFrame, ref_flux: float) -> pd.DataFrame:
    out = df.copy()
    out["oiii_ref_flux"] = ref_flux
    out["oiii_scale_factor"] = out["oiii_ref_flux"] / out["oiii5007_flux"]

    mappings = [
        ("ha_broad_flux", "ha_broad_err", "ha_broad_flux_norm", "ha_broad_err_norm"),
        ("hb_broad_flux", "hb_broad_err", "hb_broad_flux_norm", "hb_broad_err_norm"),
        ("cont5100_int_flux", "cont5100_err", "cont5100_int_flux_norm", "cont5100_err_norm"),
    ]

    for in_val, in_err, out_val, out_err in mappings:
        if in_val not in out.columns:
            out[out_val] = np.nan
            out[out_err] = np.nan
            continue

        vals = []
        errs = []
        for _, row in out.iterrows():
            v, e = normalize_one_value(
                row.get(in_val, np.nan),
                row.get(in_err, np.nan),
                row.get("oiii5007_flux", np.nan),
                row.get("oiii5007_err", np.nan),
                row.get("oiii_ref_flux", np.nan),
            )
            vals.append(v)
            errs.append(e)

        out[out_val] = vals
        out[out_err] = errs

    return out


# ============================================================
# СТАТИСТИКИ ПЕРЕМЕННОСТИ
# ============================================================

def clean_series(values, errors):
    y = np.asarray(values, dtype=float)
    e = np.asarray(errors, dtype=float)

    mask = np.isfinite(y) & np.isfinite(e) & (e > 0)
    return y[mask], e[mask]


def weighted_mean(y: np.ndarray, e: np.ndarray) -> float:
    w = 1.0 / (e ** 2)
    return float(np.sum(w * y) / np.sum(w))


def max_pairwise_significance(y: np.ndarray, e: np.ndarray) -> float:
    if len(y) < 2:
        return np.nan

    m = 0.0
    for i in range(len(y)):
        for j in range(i + 1, len(y)):
            den = np.sqrt(e[i] ** 2 + e[j] ** 2)
            if den > 0 and np.isfinite(den):
                s = abs(y[i] - y[j]) / den
                if np.isfinite(s):
                    m = max(m, s)
    return float(m)


def calc_fvar(y: np.ndarray, e: np.ndarray):
    n = len(y)
    if n < 2:
        return np.nan, np.nan, np.nan, np.nan

    mean_flux = np.mean(y)
    if not np.isfinite(mean_flux) or mean_flux == 0:
        return np.nan, np.nan, np.nan, np.nan

    s2 = np.var(y, ddof=1)
    mean_err2 = np.mean(e ** 2)
    sigma_xs2 = s2 - mean_err2
    sigma_nxs2 = sigma_xs2 / (mean_flux ** 2)

    if sigma_xs2 > 0:
        fvar = np.sqrt(sigma_xs2) / abs(mean_flux)
        err_fvar = np.sqrt(
            (np.sqrt(1.0 / (2.0 * n)) * mean_err2 / (abs(mean_flux) ** 2 * fvar)) ** 2 +
            (np.sqrt(mean_err2 / n) / abs(mean_flux)) ** 2
        )
    else:
        fvar = 0.0
        err_fvar = np.sqrt(mean_err2 / n) / abs(mean_flux)

    return float(sigma_xs2), float(sigma_nxs2), float(fvar), float(err_fvar)


def variability_stats_for_series(object_name: str, df: pd.DataFrame, value_col: str, err_col: str, series_name: str) -> dict:
    if value_col not in df.columns or err_col not in df.columns:
        return {
            "object_name": object_name,
            "series": series_name,
            "n_epochs": 0,
            "mean_flux": np.nan,
            "weighted_mean_flux": np.nan,
            "chi2_const": np.nan,
            "chi2_red": np.nan,
            "p_const": np.nan,
            "sigma_xs2": np.nan,
            "sigma_nxs2": np.nan,
            "fvar": np.nan,
            "fvar_err": np.nan,
            "rmax": np.nan,
            "max_pair_sig": np.nan,
            "variable": False,
            "variable_loose": False,
            "classification": "no_data",
            "criterion": "",
        }

    sub = df.copy()
    y, e = clean_series(sub[value_col].values, sub[err_col].values)
    n = len(y)

    if n == 0:
        return {
            "object_name": object_name,
            "series": series_name,
            "n_epochs": 0,
            "mean_flux": np.nan,
            "weighted_mean_flux": np.nan,
            "chi2_const": np.nan,
            "chi2_red": np.nan,
            "p_const": np.nan,
            "sigma_xs2": np.nan,
            "sigma_nxs2": np.nan,
            "fvar": np.nan,
            "fvar_err": np.nan,
            "rmax": np.nan,
            "max_pair_sig": np.nan,
            "variable": False,
            "variable_loose": False,
            "classification": "no_data",
            "criterion": "",
        }

    mean_flux = float(np.mean(y))
    wmean = weighted_mean(y, e) if n >= 1 else np.nan

    if n >= 2:
        chi2_val = float(np.sum(((y - wmean) / e) ** 2))
        dof = n - 1
        chi2_red = chi2_val / dof if dof > 0 else np.nan
        p_const = float(chi2_dist.sf(chi2_val, dof)) if dof > 0 else np.nan
    else:
        chi2_val = np.nan
        chi2_red = np.nan
        p_const = np.nan

    sigma_xs2, sigma_nxs2, fvar, fvar_err = calc_fvar(y, e)

    ymax = np.nanmax(y)
    ymin = np.nanmin(y)
    rmax = float(ymax / ymin) if np.isfinite(ymin) and ymin > 0 else np.nan

    pair_sig = max_pairwise_significance(y, e)

    variable = False
    variable_loose = False
    classification = "insufficient_data"
    criterion = ""

    if n == 1:
        classification = "insufficient_data"
        criterion = "N<2"
    elif n == 2:
        variable = bool(np.isfinite(pair_sig) and pair_sig >= PAIRWISE_SIG_THRESHOLD)
        variable_loose = bool(np.isfinite(pair_sig) and pair_sig >= PAIRWISE_SIG_THRESHOLD)
        classification = "variable_2epoch" if variable else "consistent_2epoch"
        criterion = f"max_pair_sig {'>=' if variable else '<'} {PAIRWISE_SIG_THRESHOLD}"
    else:
        cond_chi2 = np.isfinite(p_const) and (p_const < P_CONST_THRESHOLD)
        cond_xs = np.isfinite(sigma_xs2) and (sigma_xs2 > 0)
        cond_fvar = np.isfinite(fvar) and np.isfinite(fvar_err) and (fvar > FVAR_SIGMA_THRESHOLD * fvar_err)

        variable = bool(cond_chi2 and cond_xs and cond_fvar)
        variable_loose = bool((np.isfinite(p_const) and p_const < 0.05) and cond_xs)

        classification = "variable" if variable else "consistent_with_errors"
        criterion = f"p_const<{P_CONST_THRESHOLD}, sigma_xs2>0, Fvar>{FVAR_SIGMA_THRESHOLD}*err"

    return {
        "object_name": object_name,
        "series": series_name,
        "n_epochs": int(n),
        "mean_flux": mean_flux,
        "weighted_mean_flux": wmean,
        "chi2_const": chi2_val,
        "chi2_red": chi2_red,
        "p_const": p_const,
        "sigma_xs2": sigma_xs2,
        "sigma_nxs2": sigma_nxs2,
        "fvar": fvar,
        "fvar_err": fvar_err,
        "rmax": rmax,
        "max_pair_sig": pair_sig,
        "variable": variable,
        "variable_loose": variable_loose,
        "classification": classification,
        "criterion": criterion,
    }


# ============================================================
# ПЛОТЫ
# ============================================================

def get_survey_style(survey: str) -> dict:
    return SURVEY_STYLES.get(str(survey).upper(), SURVEY_STYLES["OTHER"])


def add_series_plot(ax, df: pd.DataFrame, value_col: str, err_col: str, title: str, color: str, stats_row: Optional[pd.Series] = None):
    sub = df.copy()
    sub = sub[np.isfinite(sub["date_mjd"])]
    if value_col in sub.columns:
        sub = sub[np.isfinite(sub[value_col])]
    else:
        sub = sub.iloc[0:0]

    if sub.empty:
        ax.text(0.5, 0.5, "Нет данных", ha="center", va="center", transform=ax.transAxes)
        ax.set_title(title)
        ax.grid(True, alpha=0.3)
        return

    sub = sub.sort_values("date_mjd")
    sub["date_dt"] = sub["date_mjd"].apply(mjd_to_datetime)

    ax.plot(sub["date_dt"], sub[value_col], color=color, lw=1.2, alpha=0.7, zorder=2)

    for survey in ["SDSS", "DESI", "OTHER"]:
        grp = sub[sub["survey"].astype(str).str.upper() == survey]
        if grp.empty:
            continue
        style = get_survey_style(survey)
        yerr = grp[err_col].fillna(0.0).to_numpy(dtype=float) if err_col in grp.columns else None
        ax.errorbar(
            grp["date_dt"],
            grp[value_col],
            yerr=yerr,
            fmt=style["marker"],
            color=style["color"],
            ecolor=style["color"],
            elinewidth=1.0,
            capsize=3,
            markersize=6,
            linestyle="none",
            alpha=0.95,
            zorder=3,
        )

    ax.set_title(title)
    ax.grid(True, linestyle="--", alpha=0.35)

    locator = mdates.AutoDateLocator(minticks=3, maxticks=6)
    formatter = mdates.ConciseDateFormatter(locator)
    ax.xaxis.set_major_locator(locator)
    ax.xaxis.set_major_formatter(formatter)

    if stats_row is not None and not pd.isna(stats_row.get("n_epochs", np.nan)):
        txt = (
            f"N={int(stats_row['n_epochs'])} | "
            f"p={stats_row['p_const']:.2e} | "
            f"Fvar={stats_row['fvar']:.3f}±{stats_row['fvar_err']:.3f} | "
            f"var={bool(stats_row['variable'])}"
        )
        ax.text(
            0.01, 0.95, txt,
            transform=ax.transAxes,
            ha="left", va="top",
            fontsize=9,
            bbox=dict(boxstyle="round", fc="white", ec="0.7", alpha=0.85),
        )


def save_variability_plot(object_dir: Path, df: pd.DataFrame, stats_df: pd.DataFrame):
    if df.empty:
        return

    use = df.copy()
    use["date_mjd"] = pd.to_numeric(use.get("date_mjd", np.nan), errors="coerce")
    use = use[np.isfinite(use["date_mjd"])].copy()
    if use.empty:
        return

    def pick_stats(series_name: str):
        sub = stats_df[stats_df["series"] == series_name]
        return sub.iloc[0] if len(sub) > 0 else None

    fig, axes = plt.subplots(4, 1, figsize=(12, 12), sharex=True)

    add_series_plot(
        axes[0], use,
        "oiii5007_flux", "oiii5007_err",
        "[O III] λ5007 raw flux (diagnostic for calibration)",
        color="dimgray",
        stats_row=pick_stats("oiii5007_raw"),
    )

    add_series_plot(
        axes[1], use,
        "ha_broad_flux_norm", "ha_broad_err_norm",
        "Hα broad flux normalized by [O III]",
        color="crimson",
        stats_row=pick_stats("ha_broad_flux_norm"),
    )

    add_series_plot(
        axes[2], use,
        "hb_broad_flux_norm", "hb_broad_err_norm",
        "Hβ broad flux normalized by [O III]",
        color="royalblue",
        stats_row=pick_stats("hb_broad_flux_norm"),
    )

    add_series_plot(
        axes[3], use,
        "cont5100_int_flux_norm", "cont5100_err_norm",
        "Continuum 5100±50 Å normalized by [O III]",
        color="seagreen",
        stats_row=pick_stats("cont5100_int_flux_norm"),
    )

    axes[-1].set_xlabel("Date")
    fig.suptitle(f"{object_dir.name}\nNormalized variability diagnostics", fontsize=14, y=0.98)

    plt.tight_layout(rect=[0.02, 0.02, 0.98, 0.96])
    fig.savefig(object_dir / OUTPUT_PLOT, dpi=160, bbox_inches="tight")
    plt.close(fig)


# ============================================================
# ОБРАБОТКА ОДНОГО ОБЪЕКТА
# ============================================================

def process_object(object_dir: Path, redshift_cache: dict[int, dict]):
    print(f"\n=== {object_dir.name} ===")

    measurement_df = build_measurement_table(object_dir)
    measurement_df = ensure_file_name_column(measurement_df)

    if measurement_df.empty:
        print("  Нет line/continuum таблиц")
        return pd.DataFrame(), pd.DataFrame(), {}

    files = list_spectrum_files(object_dir)
    if not files:
        print("  Нет FITS-спектров")
        return pd.DataFrame(), pd.DataFrame(), {}

    measurement_map = {}
    if "file_name" in measurement_df.columns:
        for _, row in measurement_df.iterrows():
            fname = row.get("file_name")
            if fname is not None and not pd.isna(fname):
                measurement_map[str(fname)] = row.to_dict()

    oiii_rows = []

    for path in files:
        print(f"  -> [O III] {path.name}")

        row_in = measurement_map.get(path.name, {})

        try:
            raw = load_raw_spectrum(path)
        except Exception as e:
            oiii_rows.append({
                "object_name": object_dir.name,
                "file_name": path.name,
                "spectrum_file": str(path),
                "survey": guess_survey_from_path(path),
                "specid": None,
                "targetid": None,
                "date_obs": "",
                "date_mjd": np.nan,
                "redshift": np.nan,
                "redshift_source": "",
                "oiii5007_flux": np.nan,
                "oiii5007_err": np.nan,
                "oiii5007_snr": np.nan,
                "oiii_method": "",
                "oiii_plot_file": "",
                "status": "failed",
                "notes": f"fits read failed: {e}",
            })
            continue

        specid = raw.get("specid")
        targetid = raw.get("targetid")

        redshift = first_existing_value(
            row_in,
            ["redshift", "redshift_line", "redshift_cont"],
            as_float=True,
        )
        redshift_source = "measurement_table"

        if not np.isfinite(redshift):
            redshift = safe_float(raw.get("redshift"))
            if np.isfinite(redshift):
                redshift_source = raw.get("redshift_source", "fits")

        if not np.isfinite(redshift) and specid is not None and specid in redshift_cache:
            redshift = safe_float(redshift_cache[specid].get("redshift"))
            if np.isfinite(redshift):
                redshift_source = "redshift_cache"

        date_obs = first_existing_value(row_in, ["date_obs", "date_obs_line", "date_obs_cont"], as_float=False)
        if date_obs is None:
            date_obs = raw.get("date_obs", "")

        date_mjd = first_existing_value(row_in, ["date_mjd", "date_mjd_line", "date_mjd_cont"], as_float=True)
        if not np.isfinite(date_mjd):
            date_mjd = safe_float(raw.get("date_mjd"))
        if not np.isfinite(date_mjd) and date_obs:
            date_mjd = parse_date_to_mjd(str(date_obs))

        if targetid is None:
            targetid = first_existing_value(row_in, ["targetid", "targetid_line", "targetid_cont"], as_float=False)
            targetid = safe_int(targetid)

        if specid is None:
            specid = first_existing_value(row_in, ["specid", "specid_line", "specid_cont"], as_float=False)
            specid = safe_int(specid)

        if not np.isfinite(redshift):
            oiii_rows.append({
                "object_name": object_dir.name,
                "file_name": path.name,
                "spectrum_file": str(path),
                "survey": raw.get("survey", guess_survey_from_path(path)),
                "specid": specid,
                "targetid": targetid,
                "date_obs": date_obs or "",
                "date_mjd": date_mjd,
                "redshift": np.nan,
                "redshift_source": "",
                "oiii5007_flux": np.nan,
                "oiii5007_err": np.nan,
                "oiii5007_snr": np.nan,
                "oiii_method": "",
                "oiii_plot_file": "",
                "status": "failed",
                "notes": "redshift not found",
            })
            continue

        try:
            meas = measure_oiii_5007(
                wave_obs=raw["wave_obs"],
                flux=raw["flux"],
                ivar=raw["ivar"],
                redshift=redshift,
            )

            flux_oiii = meas["flux_oiii"]
            err_oiii = meas["err_oiii"]
            snr_oiii = flux_oiii / err_oiii if np.isfinite(err_oiii) and err_oiii > 0 else np.nan

            plot_path = path.with_name(f"{strip_fits_suffix(path)}_oiii5007.png")
            try:
                save_oiii_plot(
                    out_path=plot_path,
                    object_name=object_dir.name,
                    file_name=path.name,
                    survey=raw.get("survey", guess_survey_from_path(path)),
                    date_obs=str(date_obs or ""),
                    redshift=redshift,
                    meas=meas,
                )
                plot_file = str(plot_path)
            except Exception:
                plot_file = ""

            oiii_rows.append({
                "object_name": object_dir.name,
                "file_name": path.name,
                "spectrum_file": str(path),
                "survey": raw.get("survey", guess_survey_from_path(path)),
                "specid": specid,
                "targetid": targetid,
                "date_obs": date_obs or "",
                "date_mjd": date_mjd,
                "redshift": redshift,
                "redshift_source": redshift_source,
                "oiii5007_flux": flux_oiii,
                "oiii5007_err": err_oiii,
                "oiii5007_snr": snr_oiii,
                "oiii_method": meas["method"],
                "oiii_plot_file": plot_file,
                "status": "ok" if np.isfinite(flux_oiii) else "failed",
                "notes": "ok",
            })

        except Exception as e:
            oiii_rows.append({
                "object_name": object_dir.name,
                "file_name": path.name,
                "spectrum_file": str(path),
                "survey": raw.get("survey", guess_survey_from_path(path)),
                "specid": specid,
                "targetid": targetid,
                "date_obs": date_obs or "",
                "date_mjd": date_mjd,
                "redshift": redshift,
                "redshift_source": redshift_source,
                "oiii5007_flux": np.nan,
                "oiii5007_err": np.nan,
                "oiii5007_snr": np.nan,
                "oiii_method": "",
                "oiii_plot_file": "",
                "status": "failed",
                "notes": f"oiii measure failed: {e}",
            })

    oiii_df = pd.DataFrame(oiii_rows)
    oiii_df["date_mjd"] = pd.to_numeric(oiii_df["date_mjd"], errors="coerce")
    oiii_df = oiii_df.sort_values(by=["date_mjd", "survey", "file_name"], ascending=[True, True, True]).reset_index(drop=True)
    save_tsv(oiii_df, object_dir / OIII_NORM_TABLE, preferred_cols=PREFERRED_OIII_COLS)

    good_ref = (
        np.isfinite(oiii_df["oiii5007_flux"]) &
        (oiii_df["oiii5007_flux"] > 0)
    )

    if "oiii5007_snr" in oiii_df.columns:
        good_ref &= (np.isfinite(oiii_df["oiii5007_snr"]) & (oiii_df["oiii5007_snr"] >= MIN_OIII_SNR_FOR_REFERENCE))

    ref_candidates = oiii_df.loc[good_ref, "oiii5007_flux"].to_numpy(dtype=float)

    if len(ref_candidates) == 0:
        fallback = oiii_df.loc[np.isfinite(oiii_df["oiii5007_flux"]) & (oiii_df["oiii5007_flux"] > 0), "oiii5007_flux"].to_numpy(dtype=float)
        ref_flux = float(np.nanmedian(fallback)) if len(fallback) > 0 else np.nan
    else:
        ref_flux = float(np.nanmedian(ref_candidates))

    merged = pd.merge(
        measurement_df,
        oiii_df[[
            "file_name",
            "oiii5007_flux",
            "oiii5007_err",
            "oiii5007_snr",
            "oiii_method",
            "status",
            "notes",
            "redshift",
            "date_obs",
            "date_mjd",
            "survey",
            "spectrum_file",
        ]],
        on="file_name",
        how="left",
        suffixes=("", "_oiii"),
    )

    for col in ["survey", "spectrum_file", "date_obs", "date_mjd", "redshift"]:
        alt = f"{col}_oiii"
        if col not in merged.columns and alt in merged.columns:
            merged[col] = merged[alt]
        elif col in merged.columns and alt in merged.columns:
            merged[col] = merged[col].combine_first(merged[alt])

    norm_df = apply_oiii_normalization(merged, ref_flux=ref_flux)
    norm_df["date_mjd"] = pd.to_numeric(norm_df.get("date_mjd", np.nan), errors="coerce")
    norm_df = norm_df.sort_values(by=["date_mjd", "survey", "file_name"], ascending=[True, True, True]).reset_index(drop=True)
    save_tsv(norm_df, object_dir / OUTPUT_NORM_TABLE, preferred_cols=PREFERRED_NORM_COLS)

    stats_rows = []
    stats_rows.append(variability_stats_for_series(object_dir.name, norm_df, "oiii5007_flux", "oiii5007_err", "oiii5007_raw"))
    stats_rows.append(variability_stats_for_series(object_dir.name, norm_df, "ha_broad_flux_norm", "ha_broad_err_norm", "ha_broad_flux_norm"))
    stats_rows.append(variability_stats_for_series(object_dir.name, norm_df, "hb_broad_flux_norm", "hb_broad_err_norm", "hb_broad_flux_norm"))
    stats_rows.append(variability_stats_for_series(object_dir.name, norm_df, "cont5100_int_flux_norm", "cont5100_err_norm", "cont5100_int_flux_norm"))

    stats_df = pd.DataFrame(stats_rows)
    save_tsv(stats_df, object_dir / OUTPUT_VAR_TABLE, preferred_cols=PREFERRED_VAR_COLS)

    try:
        save_variability_plot(object_dir, norm_df, stats_df)
    except Exception as e:
        print(f"  plot failed: {e}")

    def get_flag(series_name: str, col: str, default=np.nan):
        sub = stats_df[stats_df["series"] == series_name]
        if len(sub) == 0 or col not in sub.columns:
            return default
        return sub.iloc[0][col]

    summary = {
        "object_name": object_dir.name,
        "rm_candidate": False,
        "cont_variable": bool(get_flag("cont5100_int_flux_norm", "variable", False)),
        "ha_variable": bool(get_flag("ha_broad_flux_norm", "variable", False)),
        "hb_variable": bool(get_flag("hb_broad_flux_norm", "variable", False)),
        "n_spectra_total": int(len(norm_df)),
        "n_oiii_measured": int(np.sum(np.isfinite(norm_df["oiii5007_flux"]))),
        "oiii_ref_flux": ref_flux,
        "oiii_raw_p_const": get_flag("oiii5007_raw", "p_const"),
        "oiii_raw_fvar": get_flag("oiii5007_raw", "fvar"),
        "ha_n": get_flag("ha_broad_flux_norm", "n_epochs"),
        "ha_p_const": get_flag("ha_broad_flux_norm", "p_const"),
        "ha_fvar": get_flag("ha_broad_flux_norm", "fvar"),
        "ha_fvar_err": get_flag("ha_broad_flux_norm", "fvar_err"),
        "hb_n": get_flag("hb_broad_flux_norm", "n_epochs"),
        "hb_p_const": get_flag("hb_broad_flux_norm", "p_const"),
        "hb_fvar": get_flag("hb_broad_flux_norm", "fvar"),
        "hb_fvar_err": get_flag("hb_broad_flux_norm", "fvar_err"),
        "cont_n": get_flag("cont5100_int_flux_norm", "n_epochs"),
        "cont_p_const": get_flag("cont5100_int_flux_norm", "p_const"),
        "cont_fvar": get_flag("cont5100_int_flux_norm", "fvar"),
        "cont_fvar_err": get_flag("cont5100_int_flux_norm", "fvar_err"),
    }

    summary["rm_candidate"] = bool(summary["cont_variable"] and (summary["ha_variable"] or summary["hb_variable"]))

    return norm_df, stats_df, summary


# ============================================================
# MAIN
# ============================================================

def main():
    if not ROOT_DIR.exists():
        raise FileNotFoundError(f"Не найдена директория: {ROOT_DIR}")

    redshift_cache = load_redshift_cache(ROOT_DIR)
    object_dirs = list_object_dirs(ROOT_DIR)

    if not object_dirs:
        print("Папки объектов не найдены.")
        return

    all_stats = []
    all_summary = []

    for object_dir in object_dirs:
        try:
            _, stats_df, summary = process_object(object_dir, redshift_cache)
            if not stats_df.empty:
                all_stats.append(stats_df)
            if summary:
                all_summary.append(summary)
        except Exception as e:
            print(f"[ERR] {object_dir.name}: {e}")

    if all_stats:
        master_stats = pd.concat(all_stats, ignore_index=True)
    else:
        master_stats = pd.DataFrame(columns=PREFERRED_VAR_COLS)

    summary_df = pd.DataFrame(all_summary)
    if not summary_df.empty:
        summary_df = summary_df.sort_values(
            by=["rm_candidate", "cont_variable", "ha_variable", "hb_variable", "object_name"],
            ascending=[False, False, False, False, True]
        ).reset_index(drop=True)

    save_table_readable(
        master_stats,
        ROOT_DIR / MASTER_VAR_TABLE_TSV,
        ROOT_DIR / MASTER_VAR_TABLE_XLSX,
        preferred_cols=PREFERRED_VAR_COLS,
    )

    save_table_readable(
        summary_df,
        ROOT_DIR / MASTER_OBJ_TABLE_TSV,
        ROOT_DIR / MASTER_OBJ_TABLE_XLSX,
        preferred_cols=PREFERRED_SUMMARY_COLS,
    )

    print("\nГотово.")
    print(f"Общие статистики TSV:   {ROOT_DIR / MASTER_VAR_TABLE_TSV}")
    print(f"Общие статистики XLSX:  {ROOT_DIR / MASTER_VAR_TABLE_XLSX}")
    print(f"Сводка кандидатов TSV:  {ROOT_DIR / MASTER_OBJ_TABLE_TSV}")
    print(f"Сводка кандидатов XLSX: {ROOT_DIR / MASTER_OBJ_TABLE_XLSX}")


if __name__ == "__main__":
    main()